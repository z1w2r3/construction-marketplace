"""
Excel 文档解析器模块

解析 .xlsx 和 .xls 格式的 Excel 文档
提取工作表、单元格数据、公式等信息
"""
from typing import Dict, List, Optional, Any
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from .base_parser import BaseParser
from utils import get_logger, ParseError

logger = get_logger(__name__)


class ExcelParser(BaseParser):
    """Excel 文档解析器"""

    def __init__(self):
        super().__init__()

    def get_supported_extensions(self) -> list:
        """获取支持的文件扩展名"""
        return ['.xlsx', '.xls']

    def parse(self, file_path: str, options: Optional[Dict[str, Any]] = None) -> Dict:
        """
        解析 Excel 文档

        Args:
            file_path: Excel 文档路径
            options: 解析选项
                - sheet_name: 指定工作表名称
                - max_rows: 每个工作表最大行数
                - max_sheets: 最大工作表数

        Returns:
            解析结果字典
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ParseError(
                "缺少 openpyxl 库，请运行: pip install openpyxl"
            )

        options = options or {}
        sheet_name = options.get('sheet_name')
        max_rows = options.get('max_rows', 100)
        max_sheets = options.get('max_sheets', 10)

        try:
            # 加载工作簿 (只读模式提高性能)
            self.logger.info(f"加载 Excel 文档: {file_path}")
            wb = load_workbook(file_path, read_only=True, data_only=True)

            # 提取内容
            content = {}

            # 1. 获取工作表列表
            sheet_names = wb.sheetnames
            content['sheet_names'] = sheet_names

            # 2. 提取工作表数据
            sheets_to_read = [sheet_name] if sheet_name else sheet_names[:max_sheets]
            sheets_data = []

            for name in sheets_to_read:
                ws = wb[name]
                sheet_data = self._extract_sheet_data(ws, name, max_rows)
                sheets_data.append(sheet_data)

            content['sheets'] = sheets_data

            # 3. 生成摘要
            summary = self._generate_summary(content)

            # 4. 元数据
            metadata = self._extract_metadata(wb)

            return self._create_success_response(
                file_path,
                content,
                summary,
                metadata
            )

        except Exception as e:
            self.logger.error(f"Excel 文档解析失败: {e}", exc_info=True)
            raise ParseError(f"Excel 文档解析失败: {str(e)}")

    def _extract_sheet_data(self, ws, sheet_name: str, max_rows: int) -> Dict:
        """
        提取工作表数据

        Args:
            ws: Worksheet 对象
            sheet_name: 工作表名称
            max_rows: 最大行数

        Returns:
            工作表数据字典
        """
        sheet_data = {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "data": [],
            "headers": []
        }

        # 提取数据
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= max_rows:
                break

            # 过滤完全空的行
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # 转换为字符串列表
            row_data = [str(cell) if cell is not None else '' for cell in row]
            sheet_data["data"].append(row_data)
            row_count += 1

        # 识别表头 (第一行)
        if sheet_data["data"]:
            sheet_data["headers"] = sheet_data["data"][0]

        self.logger.info(
            f"提取工作表 '{sheet_name}': "
            f"{len(sheet_data['data'])} 行 x {ws.max_column} 列"
        )

        return sheet_data

    def _extract_metadata(self, wb) -> Dict:
        """提取元数据"""
        metadata = {}

        try:
            props = wb.properties

            if hasattr(props, 'creator') and props.creator:
                metadata['creator'] = props.creator

            if hasattr(props, 'title') and props.title:
                metadata['title'] = props.title

            if hasattr(props, 'created') and props.created:
                metadata['created'] = str(props.created)

            if hasattr(props, 'modified') and props.modified:
                metadata['modified'] = str(props.modified)

        except Exception as e:
            self.logger.warning(f"提取元数据失败: {e}")

        return metadata

    def _generate_summary(self, content: Dict) -> Dict:
        """生成摘要"""
        total_rows = sum(len(sheet['data']) for sheet in content['sheets'])
        total_sheets = len(content['sheets'])

        summary = {
            "total_sheets": total_sheets,
            "total_rows": total_rows,
            "sheet_names": content['sheet_names']
        }

        return summary
